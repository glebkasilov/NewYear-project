import openpyxl

from PyQt6 import uic, QtGui
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QApplication, QMainWindow, QDialog, QFileDialog, QMessageBox

from app.database.repository import GiftsRepository


class ExportExcel(QDialog):
    def __init__(self, email):
        super().__init__()
        uic.loadUi('app/style/dialog/export_xlxs.ui', self)
        self.email = email
        self.folder_path = ""
        self.show()

        self.routeButton.clicked.connect(self.select_folder)
        self.saveButton.clicked.connect(self.export)
        self.cancelButton.clicked.connect(self.close)

    def select_folder(self):
        folder_path = QFileDialog.getExistingDirectory(self, "Выберите папку")
        self.folder_path = folder_path

        if folder_path:
            self.routeEdit.setText(f"Выбранная папка: {folder_path}")
        else:
            self.routeEdit.setText("Папка не выбрана.")

    def show_error_message(self, message="Внутряняя ошибка"):
        msg_box = QMessageBox(self)
        msg_box.setIcon(QMessageBox.Icon.Critical)
        msg_box.setWindowTitle("Ошибка")
        msg_box.setText(message)
        msg_box.setStandardButtons(QMessageBox.StandardButton.Cancel)

        msg_box.exec()

    def export(self):
        if self.folder_path == "":
            self.show_error_message("Путь не указан")
            return

        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = 'Название'
        ws['B1'] = 'Количество'
        ws['C1'] = 'Цена за единицу'
        ws['D1'] = 'Цена'
        ws['E1'] = 'Описание'
        ws['F1'] = 'Готовность'

        gifts = GiftsRepository.get_list(self.email)

        if not gifts:
            self.show_error_message("Список пуст")
            return

        for i, gift in enumerate(gifts, start=2):
            ws.cell(row=i, column=1).value = gift[0]
            ws.cell(row=i, column=2).value = gift[1]
            ws.cell(row=i, column=3).value = gift[2]
            ws.cell(row=i, column=4).value = gift[1] * gift[2]
            ws.cell(row=i, column=5).value = gift[3]
            ws.cell(
                row=i, column=6).value = "Купленно" if gift[4] else "Не купленно"

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 15

        wb.save(self.folder_path + '/' +
                "Список подарков_" + f'{self.email}.xlsx')
        self.close()
